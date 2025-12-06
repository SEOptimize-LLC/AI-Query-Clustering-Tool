"""
Excel export functionality for clustering results.
Uses openpyxl for rich Excel formatting.
"""
import io
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference


class ExcelExporter:
    """Exports clustering results to Excel with formatting."""
    
    # Style definitions
    HEADER_FILL = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid"
    )
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    
    INTENT_COLORS = {
        "informational": "92D050",  # Green
        "transactional": "FF6B6B",  # Red
        "commercial": "FFB347",     # Orange
        "navigational": "6BB3FF",   # Blue
        "unknown": "CCCCCC"         # Gray
    }
    
    def __init__(self):
        """Initialize exporter."""
        self.wb = None
    
    def export_full_report(
        self,
        job_data: Dict[str, Any]
    ) -> bytes:
        """
        Export complete report with multiple sheets.
        
        Args:
            job_data: Complete job data
        
        Returns:
            Excel file as bytes
        """
        self.wb = Workbook()
        
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Create sheets
        self._create_summary_sheet(job_data)
        self._create_clusters_sheet(job_data.get("clusters", []))
        self._create_keywords_sheet(job_data.get("clusters", []))
        
        unclustered = job_data.get("unclustered", [])
        if unclustered:
            self._create_unclustered_sheet(unclustered)
        
        # Save to bytes
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _create_summary_sheet(self, job_data: Dict[str, Any]):
        """Create summary sheet with overview metrics."""
        ws = self.wb.create_sheet("Summary")
        
        # Title
        ws["A1"] = "Keyword Clustering Report"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")
        
        # Metadata
        ws["A3"] = "Generated:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Metrics
        metrics = [
            ("Total Keywords", job_data.get("total_keywords", 0)),
            ("Total Clusters", job_data.get("total_clusters", 0)),
            ("Cluster Rate", f"{job_data.get('cluster_rate', 0):.1%}"),
            ("Total Search Volume", f"{job_data.get('total_volume', 0):,}"),
            ("Average Difficulty", f"{job_data.get('avg_difficulty', 0):.1f}"),
            ("Unclustered Keywords", job_data.get("unclustered_count", 0)),
        ]
        
        row = 5
        for label, value in metrics:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1
        
        # Intent distribution
        row += 2
        ws[f"A{row}"] = "Intent Distribution"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1
        
        intent_counts = job_data.get("intent_distribution", {})
        for intent, count in intent_counts.items():
            ws[f"A{row}"] = intent.title()
            ws[f"B{row}"] = count
            
            # Color cell
            color = self.INTENT_COLORS.get(intent.lower(), "CCCCCC")
            ws[f"A{row}"].fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"
            )
            row += 1
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
    
    def _create_clusters_sheet(self, clusters: List[Dict]):
        """Create clusters overview sheet."""
        ws = self.wb.create_sheet("Clusters")
        
        # Headers
        headers = [
            "Cluster ID",
            "Label",
            "Size",
            "Total Volume",
            "Avg. Difficulty",
            "Intent",
            "Quality Score",
            "Top Keywords"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, cluster in enumerate(clusters, 2):
            cluster_id = cluster.get("id", 0)
            keywords = cluster.get("keywords", [])
            
            # Get top keywords
            top_kw = []
            for kw in keywords[:5]:
                if isinstance(kw, dict):
                    top_kw.append(kw.get("keyword", ""))
                else:
                    top_kw.append(str(kw))
            
            values = [
                cluster_id,
                cluster.get("label", f"Cluster {cluster_id}"),
                cluster.get("size", len(keywords)),
                cluster.get("total_volume", 0),
                round(cluster.get("avg_difficulty", 0), 1),
                cluster.get("intent", "unknown").title(),
                round(cluster.get("quality_score", 0), 1),
                ", ".join(top_kw)
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                
                # Intent color
                if col == 6:
                    intent = str(value).lower()
                    color = self.INTENT_COLORS.get(intent, "CCCCCC")
                    cell.fill = PatternFill(
                        start_color=color,
                        end_color=color,
                        fill_type="solid"
                    )
                
                # Quality color
                if col == 7:
                    score = float(value) if value else 0
                    if score >= 70:
                        color = "92D050"
                    elif score >= 40:
                        color = "FFB347"
                    else:
                        color = "FF6B6B"
                    cell.fill = PatternFill(
                        start_color=color,
                        end_color=color,
                        fill_type="solid"
                    )
        
        # Adjust column widths
        widths = [10, 30, 8, 15, 12, 15, 12, 50]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _create_keywords_sheet(self, clusters: List[Dict]):
        """Create detailed keywords sheet."""
        ws = self.wb.create_sheet("Keywords")
        
        # Headers
        headers = [
            "Keyword",
            "Cluster ID",
            "Cluster Label",
            "Search Volume",
            "Keyword Difficulty",
            "Intent"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        # Data rows
        row_idx = 2
        for cluster in clusters:
            cluster_id = cluster.get("id", 0)
            label = cluster.get("label", f"Cluster {cluster_id}")
            intent = cluster.get("intent", "unknown")
            keywords = cluster.get("keywords", [])
            
            for kw_data in keywords:
                if isinstance(kw_data, dict):
                    keyword = kw_data.get("keyword", "")
                    volume = kw_data.get("search_volume", 0)
                    difficulty = kw_data.get("keyword_difficulty", 0)
                else:
                    keyword = str(kw_data)
                    volume = 0
                    difficulty = 0
                
                values = [
                    keyword,
                    cluster_id,
                    label,
                    volume,
                    difficulty,
                    intent.title()
                ]
                
                for col, value in enumerate(values, 1):
                    ws.cell(row=row_idx, column=col, value=value)
                
                row_idx += 1
        
        # Adjust column widths
        widths = [40, 10, 30, 15, 15, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = width
        
        # Freeze header
        ws.freeze_panes = "A2"
    
    def _create_unclustered_sheet(self, keywords: List[Dict]):
        """Create sheet for unclustered keywords."""
        ws = self.wb.create_sheet("Unclustered")
        
        # Headers
        headers = ["Keyword", "Search Volume", "Keyword Difficulty"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(
                start_color="FF6B6B",
                end_color="FF6B6B",
                fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        for row_idx, kw_data in enumerate(keywords, 2):
            if isinstance(kw_data, dict):
                values = [
                    kw_data.get("keyword", ""),
                    kw_data.get("search_volume", 0),
                    kw_data.get("keyword_difficulty", 0)
                ]
            else:
                values = [str(kw_data), 0, 0]
            
            for col, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col, value=value)
        
        # Widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 18
        
        ws.freeze_panes = "A2"
    
    def export_clusters_only(
        self,
        clusters: List[Dict[str, Any]]
    ) -> bytes:
        """
        Export just clusters sheet.
        
        Args:
            clusters: List of cluster data
        
        Returns:
            Excel file bytes
        """
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        self._create_clusters_sheet(clusters)
        
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output.getvalue()